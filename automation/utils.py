import openpyxl


spreadsheet = openpyxl.load_workbook("data/data_real_state.xlsx")
main_sheet = spreadsheet['data']

def get_unique_collection_by_column(key_column):
    collection = []
    for row_index in range(2, main_sheet.max_row + 1):
        cell = main_sheet.cell(row_index, key_column)

        if cell.value not in collection:
            collection.append(cell.value)

    return collection

def count_items_by_key(key_column, key_collection):
    item_per_key = {}
    for key in key_collection:
        item_per_key[str(key)] = 0

    for row_index in range(2, main_sheet.max_row + 1):
        cell = main_sheet.cell(row_index, key_column)
        if cell.value in key_collection:
            item_per_key[str(cell.value)] = item_per_key[str(cell.value)] + 1

    return item_per_key

def sum_items_by_key(key_column, key_collection, sum_column):
    sum_per_key = {}
    for key in key_collection:
        sum_per_key[str(key)] = 0.0
    
    for row_index in range(2, main_sheet.max_row + 1):
        key_cell = main_sheet.cell(row_index, key_column)
        sum_cell = main_sheet.cell(row_index, sum_column)
        if key_cell.value in key_collection:
            sum_per_key[str(key_cell.value)] = sum_per_key[str(key_cell.value)] + (sum_cell.value)

    return sum_per_key